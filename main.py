import os
import re
import time
import imaplib
import email
from io import BytesIO
import pandas as pd
from pypdf import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from dotenv import load_dotenv
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

# Load email credentials from .env
load_dotenv()

logging.basicConfig(
    filename="app.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

EMAIL = os.getenv("EMAIL")
APP_PASSWORD = os.getenv("APP_PASSWORD")
EMAIL_SENDER = os.getenv("EMAIL_SENDER")

LAST_UID_FILE = "last_uid.txt"
REPORT_FILE = "report.xlsx"
COLUMNS = ["Date", "Visitors", "Sales"]

HEADER_COLOR = "1F4E78"
WHITE = "FFFFFF"
BORDER_COLOR = "D9D9D9"


def connect_to_email() -> imaplib.IMAP4_SSL:
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(EMAIL, APP_PASSWORD)
    return mail


def load_last_uid() -> int:
    try:
        with open(LAST_UID_FILE, "r") as f:
            return int(f.read().strip())
    except FileNotFoundError:
        return -1


def save_last_uid(uid: int) -> None:
    with open(LAST_UID_FILE, "w") as f:
        f.write(str(uid))


def create_report_file() -> None:
    if not os.path.exists(REPORT_FILE):
        pd.DataFrame(columns=COLUMNS).to_excel(
            REPORT_FILE,
            index=False
        )


def extract_metrics_from_text(text: str) -> dict | None:
    visitors = re.search(r"Visitors:\s+(\d+)", text)
    sales = re.search(r"Sales:\s+(\$?\d+\.\d+)", text)
    date = re.search(r"Date:\s+(\d{4}-\d{2}-\d{2})", text)

    if visitors and sales and date:
        return {
            "Date": date.group(1),
            "Visitors": int(visitors.group(1)),
            "Sales": float(sales.group(1).replace("$",""))
        }
    return None


def extract_from_pdf(file_bytes: bytes) -> dict | None:
    reader = PdfReader(BytesIO(file_bytes))
    text = ""
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text
    return extract_metrics_from_text(text)


def extract_from_excel(file_bytes: bytes) -> dict:
    df = pd.read_excel(BytesIO(file_bytes))
    print("date " + str(df["Date"].iloc[0]).split()[0])
    return {
        "Date": str(df["Date"].iloc[0]).split()[0],
        "Visitors": int(df["Visitors"].iloc[0]),
        "Sales": float(df["Sales"].iloc[0])
    }


def extract_from_csv(file_bytes: bytes) -> dict:
    df = pd.read_csv(BytesIO(file_bytes))
    return {
        "Date": str(df["Date"].iloc[0]),
        "Visitors": int(df["Visitors"].iloc[0]),
        "Sales": float(df["Sales"].iloc[0])
    }


def process_email(msg: email.message.Message) -> list[dict]:
    rows = []
    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename()

        try:
            if content_type == "text/plain" and not filename:
                result = extract_metrics_from_text(
                    part.get_payload(decode=True).decode(errors="ignore")
                )
                if result:
                    rows.append(result)

            elif filename and filename.lower().endswith(".pdf"):
                result = extract_from_pdf(part.get_payload(decode=True))
                if result:
                    rows.append(result)

            elif filename and filename.lower().endswith((".xlsx",".xls")):
                rows.append(extract_from_excel(part.get_payload(decode=True)))

            elif filename and filename.lower().endswith(".csv"):
                rows.append(extract_from_csv(part.get_payload(decode=True)))

        except Exception:
            logging.exception("Attachment processing error")

    return rows


def fetch_new_reports(mail: imaplib.IMAP4_SSL, last_uid: int) -> tuple[list[dict], int]:
    mail.select("INBOX")
    status, data = mail.uid("search", None, f'FROM "{EMAIL_SENDER}"')

    report_rows = []
    newest_uid = last_uid
    uids = data[0].split()

    for uid in uids:
        uid_int = int(uid)

        # Process only new emails
        if uid_int > last_uid:
            try:
                status, msg_data = mail.uid("fetch", str(uid_int), "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])

                report_rows.extend(process_email(msg))
                newest_uid = uid_int

            except Exception:
                logging.exception("Email processing error")
        elif uid == last_uid and int(max(uids)) == uid:
            break

    return report_rows, newest_uid


def update_report(new_rows: list[dict]) -> None:
    if not new_rows:
        return

    try:
        existing = pd.read_excel(REPORT_FILE)
        if not existing.empty:
            # Remove the previous summary row before recalculating totals
            existing = existing[existing["Date"].notna() & (existing["Date"] != "TOTAL:")]
    except FileNotFoundError:
        existing = pd.DataFrame(columns=COLUMNS)
    except Exception:
        logging.exception("Failed to load existing report")
        raise

    updated = pd.concat([existing, pd.DataFrame(new_rows)], ignore_index=True)

    # Prevent duplicate reports of the same date from being added twice
    updated = updated.drop_duplicates(subset=["Date"])

    # Keep TOTAL outside the Excel table so sorting does not affect it
    blank_row = pd.DataFrame([{
        "Date": "",
        "Visitors": "",
        "Sales": ""
    }])
    
    total_row = pd.DataFrame([{
        "Date":"TOTAL:",
        "Visitors":updated["Visitors"].sum(),
        "Sales":updated["Sales"].sum()
    }])

    final_df = pd.concat([updated, blank_row, total_row], ignore_index=True)
    final_df.to_excel(REPORT_FILE,index=False)


def format_report() -> None:
    wb = load_workbook(REPORT_FILE)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)

    header_font = Font(color=WHITE, bold=True, size=24)
    body_font = Font(size=16)

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR)
    )

    # Header formatting
    ws.row_dimensions[1].height = 35

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Last row is TOTAL, row before it is blank
    total_row = ws.max_row
    last_data_row = total_row - 2

    # Body formatting (exclude blank row and total)
    for row_num in range(2, last_data_row + 1):
        for cell in ws[row_num]:
            cell.font = body_font
            cell.border = thin_border

        ws.row_dimensions[row_num].height = 28

    # TOTAL row formatting
    ws.row_dimensions[total_row].height = 35

    for cell in ws[total_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Alignment
    for row in range(2, last_data_row + 1):
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        ws[f"C{row}"].alignment = Alignment(horizontal="right")

    # Column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col[0].column_letter].width = max(max_len * 3, 10)

    # Keep headers visible while scrolling
    ws.freeze_panes = "A2"

    # Excel Table (handles filter and stripes automatically)
    tab = Table(
        displayName="SalesData",
        ref=f"A1:C{last_data_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(REPORT_FILE)


def main() -> None:
    create_report_file()
    last_uid = load_last_uid()
    mail = connect_to_email()

    # Monitor the inbox for new reports
    try:
        rows, last_uid = fetch_new_reports(mail, last_uid)

        if rows:
            update_report(rows)
            format_report()

        save_last_uid(last_uid)

    except Exception:
        logging.exception("Main function error")


if __name__ == "__main__":
    main()
