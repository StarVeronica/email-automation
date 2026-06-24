import os
import re
import time
import imaplib
import email
from io import BytesIO
import pandas as pd
from pypdf import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from dotenv import load_dotenv

load_dotenv()

EMAIL = os.getenv("EMAIL")
APP_PASSWORD = os.getenv("APP_PASSWORD")
EMAIL_SENDER = os.getenv("EMAIL_SENDER")

LAST_UID_FILE = "last_uid.txt"
REPORT_FILE = "report.xlsx"
CHECK_INTERVAL = 60

HEADER_COLOR = "1F4E78"
ALT_ROW_COLOR = "F5F9FF"
WHITE = "FFFFFF"
BORDER_COLOR = "D9D9D9"

def connect_to_email():
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(EMAIL, APP_PASSWORD)
    return mail

def load_last_uid():
    try:
        with open(LAST_UID_FILE, "r") as f:
            return int(f.read().strip())
    except FileNotFoundError:
        return 0

def save_last_uid(uid):
    with open(LAST_UID_FILE, "w") as f:
        f.write(str(uid))

def create_report_file():
    if not os.path.exists(REPORT_FILE):
        pd.DataFrame(columns=["Date","Visitors","Sales"]).to_excel(REPORT_FILE,index=False)

def extract_metrics_from_text(text):
    visitors = re.search(r"Visitors:\s+(\d+)", text)
    sales = re.search(r"Sales:\s+(\$?\d+\.\d+)", text)
    date = re.search(r"Date:\s+(\d{4}-\d{2}-\d{2})", text)

    if visitors and sales and date:
        return {
            "Date": date.group(1),
            "Visitors": int(visitors.group(1)),
            "Sales": float(sales.group(1).replace("$",""))
        }
    return None

def extract_from_pdf(file_bytes):
    reader = PdfReader(BytesIO(file_bytes))
    text = ""
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text
    return extract_metrics_from_text(text)

def extract_from_excel(file_bytes):
    df = pd.read_excel(BytesIO(file_bytes))
    return {
        "Date": str(df["Date"].iloc[0]),
        "Visitors": int(df["Visitors"].iloc[0]),
        "Sales": float(df["Sales"].iloc[0])
    }

def extract_from_csv(file_bytes):
    df = pd.read_csv(BytesIO(file_bytes))
    return {
        "Date": str(df["Date"].iloc[0]),
        "Visitors": int(df["Visitors"].iloc[0]),
        "Sales": float(df["Sales"].iloc[0])
    }

def process_email(msg):
    rows = []
    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename()

        try:
            if content_type == "text/plain" and not filename:
                result = extract_metrics_from_text(
                    part.get_payload(decode=True).decode(errors="ignore")
                )
                if result:
                    rows.append(result)

            elif filename and filename.lower().endswith(".pdf"):
                result = extract_from_pdf(part.get_payload(decode=True))
                if result:
                    rows.append(result)

            elif filename and filename.lower().endswith((".xlsx",".xls")):
                rows.append(extract_from_excel(part.get_payload(decode=True)))

            elif filename and filename.lower().endswith(".csv"):
                rows.append(extract_from_csv(part.get_payload(decode=True)))

        except Exception as e:
            print(f"Attachment processing error: {e}")

    return rows

def fetch_new_reports(mail, last_uid):
    mail.select("INBOX")
    _, data = mail.uid("search", None, 'FROM "EMAIL_SENDER"')

    report_rows = []
    newest_uid = last_uid

    for uid in data[0].split():
        uid_int = int(uid)

        if uid_int <= last_uid:
            continue

        try:
            _, msg_data = mail.uid("fetch", str(uid_int), "(RFC822)")
            msg = email.message_from_bytes(msg_data[0][1])

            report_rows.extend(process_email(msg))
            newest_uid = uid_int

        except Exception as e:
            print(f"Email processing error: {e}")

    return report_rows, newest_uid

def update_report(new_rows):
    if not new_rows:
        return

    try:
        existing = pd.read_excel(REPORT_FILE)
        if not existing.empty:
            existing = existing.iloc[:-1]
    except Exception:
        existing = pd.DataFrame(columns=["Date","Visitors","Sales"])

    updated = pd.concat([existing, pd.DataFrame(new_rows)], ignore_index=True)

    total_row = pd.DataFrame([{
        "Date":"TOTAL",
        "Visitors":updated["Visitors"].sum(),
        "Sales":updated["Sales"].sum()
    }])

    final_df = pd.concat([updated,total_row], ignore_index=True)
    final_df.to_excel(REPORT_FILE,index=False)

def format_report():
    wb = load_workbook(REPORT_FILE)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)

    header_font = Font(color=WHITE, bold=True, size=24)
    body_font = Font(size=16)

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR)
    )

    ws.row_dimensions[1].height = 35

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row

    for row_num in range(2, last_row):
        fill = alt_fill if row_num % 2 == 0 else None

        for cell in ws[row_num]:
            if fill:
                cell.fill = fill
            cell.font = body_font
            cell.border = thin_border

    for cell in ws[last_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(REPORT_FILE)

def main():
    create_report_file()
    last_uid = load_last_uid()
    mail = connect_to_email()

    while True:
        try:
            rows, last_uid = fetch_new_reports(mail, last_uid)

            if rows:
                update_report(rows)
                format_report()

            save_last_uid(last_uid)

        except Exception as e:
            print(f"Main loop error: {e}")

        time.sleep(CHECK_INTERVAL)

if __name__ == "__main__":
    main()
